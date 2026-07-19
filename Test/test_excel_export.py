# -*- coding: utf-8 -*-
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from Function.Output.excel_export import export_to_excel


class _StubDataSource:
    def __init__(self, mapping):
        self.mapping = mapping

    def execute_query(self, sql: str):
        table = sql.split('"')[1]
        value = self.mapping.get(table)
        if isinstance(value, Exception):
            raise value
        if value is None:
            return None, "missing"
        return value.copy(), ""


class TestExcelExport(unittest.TestCase):
    def test_export_single_table_creates_visible_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            filepath = Path(tmp) / 'single.xlsx'
            datasource = _StubDataSource({
                'src1__Sheet1': pd.DataFrame({'name': ['A', 'B'], 'value': [1, 2]}),
            })

            result = export_to_excel(datasource, ['src1__Sheet1'], str(filepath))

            self.assertEqual(result['written_count'], 1)
            self.assertEqual(result['exported_tables'], ['src1__Sheet1'])
            workbook = load_workbook(filepath)
            try:
                self.assertEqual(workbook.sheetnames, ['src1__Sheet1'])
                self.assertEqual(workbook.active.title, 'src1__Sheet1')
                visible = [ws.title for ws in workbook.worksheets if ws.sheet_state == 'visible']
                self.assertEqual(visible, ['src1__Sheet1'])
            finally:
                workbook.close()

    def test_export_rejects_empty_workbook_with_friendly_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            filepath = Path(tmp) / 'empty.xlsx'
            datasource = _StubDataSource({
                'src1__Sheet1': pd.DataFrame(),
            })

            with self.assertRaisesRegex(ValueError, '暂无可导出的数据'):
                export_to_excel(datasource, ['src1__Sheet1'], str(filepath))

            self.assertFalse(filepath.exists())

    def test_export_sanitizes_and_deduplicates_sheet_names(self):
        with tempfile.TemporaryDirectory() as tmp:
            filepath = Path(tmp) / 'sanitized.xlsx'
            source_table = 'bad/name*with?chars[demo]and_a_very_long_suffix_001'
            duplicate_table = 'bad\\name*with?chars[demo]and_a_very_long_suffix_001'
            datasource = _StubDataSource({
                source_table: pd.DataFrame({'x': [1]}),
                duplicate_table: pd.DataFrame({'x': [2]}),
            })

            result = export_to_excel(datasource, [source_table, duplicate_table], str(filepath))

            self.assertEqual(result['written_count'], 2)
            workbook = load_workbook(filepath)
            try:
                self.assertEqual(len(workbook.sheetnames), 2)
                self.assertEqual(len(set(workbook.sheetnames)), 2)
                for name in workbook.sheetnames:
                    self.assertTrue(name)
                    self.assertLessEqual(len(name), 31)
                    self.assertIsNone(__import__('re').search(r'[\\/*?:\[\]]', name))
                self.assertEqual(workbook.active.sheet_state, 'visible')
            finally:
                workbook.close()


if __name__ == '__main__':
    unittest.main()
